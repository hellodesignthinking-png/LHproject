"""
ZeroSite v4.0 Excel Comparison Report Generator
================================================

다중 부지 비교 엑셀 보고서 생성

Author: ZeroSite Reporting Team
Date: 2025-12-27
Version: 1.0
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime


class ExcelComparisonReportGenerator:
    """엑셀 비교 보고서 생성기"""
    
    def __init__(self, output_dir: str = "output/comparison"):
        """
        Args:
            output_dir: 엑셀 파일 저장 디렉터리
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # 스타일 정의
        self.header_font = Font(name='맑은 고딕', size=12, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.title_font = Font(name='맑은 고딕', size=14, bold=True)
        self.normal_font = Font(name='맑은 고딕', size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_comparison_report(
        self,
        comparison_result: Dict[str, Any],
        file_name: str = None
    ) -> str:
        """
        다중 부지 비교 보고서 생성
        
        Args:
            comparison_result: M8 비교 분석 결과
            file_name: 저장 파일명 (None이면 자동 생성)
            
        Returns:
            저장된 파일 경로
        """
        # 워크북 생성
        wb = Workbook()
        
        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 1. 종합 요약 시트
        self._create_summary_sheet(wb, comparison_result)
        
        # 2. 상세 비교 시트
        self._create_detailed_comparison_sheet(wb, comparison_result)
        
        # 3. 재무 분석 시트
        self._create_financial_analysis_sheet(wb, comparison_result)
        
        # 4. LH 평가 시트
        self._create_lh_evaluation_sheet(wb, comparison_result)
        
        # 5. 추천 순위 시트
        self._create_ranking_sheet(wb, comparison_result)
        
        # 파일명 생성
        if not file_name:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"Comparison_Report_{timestamp}.xlsx"
        
        # 저장
        output_path = self.output_dir / file_name
        wb.save(str(output_path))
        
        return str(output_path)
    
    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """종합 요약 시트 생성"""
        ws = wb.create_sheet("종합 요약", 0)
        
        # 제목
        ws.merge_cells('A1:F1')
        ws['A1'] = f"다중 부지 비교 분석 보고서"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 보고서 정보
        ws['A3'] = "보고서 ID:"
        ws['B3'] = data.get("report_id", "N/A")
        ws['A4'] = "생성 일시:"
        ws['B4'] = data.get("generated_at", "N/A")
        ws['A5'] = "분석 부지 수:"
        ws['B5'] = len(data.get("sites_analyzed", []))
        
        # 통계 요약
        stats = data.get("summary_statistics", {})
        
        ws['A7'] = "평균 LH 점수:"
        ws['B7'] = f"{stats.get('average_lh_score', 0):.1f}/100"
        ws['A8'] = "평균 NPV:"
        ws['B8'] = f"₩{self._format_currency(stats.get('average_npv', 0))}"
        ws['A9'] = "평균 IRR:"
        ws['B9'] = f"{stats.get('average_irr', 0):.2f}%"
        
        # GO 판정 통계
        ws['A11'] = "판정 통계:"
        ws['A12'] = "GO:"
        ws['B12'] = stats.get('go_count', 0)
        ws['A13'] = "조건부 GO:"
        ws['B13'] = stats.get('conditional_go_count', 0)
        ws['A14'] = "NO_GO:"
        ws['B14'] = stats.get('no_go_count', 0)
        
        # 최고 부지
        best_site = data.get("best_site")
        if best_site:
            ws['D7'] = "최고 부지:"
            ws['E7'] = best_site.get("site_id", "N/A")
            ws['D8'] = "주소:"
            ws['E8'] = best_site.get("address", "N/A")
            ws['D9'] = "LH 점수:"
            ws['E9'] = f"{best_site.get('lh_score', 0):.1f}/100"
        
        # 스타일 적용
        for row in ws['A1:F14']:
            for cell in row:
                cell.font = self.normal_font
                cell.border = self.border
    
    def _create_detailed_comparison_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """상세 비교 시트 생성"""
        ws = wb.create_sheet("상세 비교")
        
        # 헤더
        headers = [
            "순위", "부지 ID", "주소", "면적(㎡)", "용도지역",
            "LH 점수", "판정", "등급", "NPV", "IRR(%)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # 데이터 입력
        sites = data.get("sites_analyzed", [])
        
        for row_idx, site in enumerate(sites, 2):
            land_info = site.get("land_info", {})
            lh_review = site.get("lh_review", {})
            feasibility = site.get("feasibility", {})
            
            ws.cell(row=row_idx, column=1, value=site.get("rank", row_idx-1))
            ws.cell(row=row_idx, column=2, value=site.get("site_id", "N/A"))
            ws.cell(row=row_idx, column=3, value=land_info.get("address", "N/A"))
            ws.cell(row=row_idx, column=4, value=land_info.get("area_sqm", 0))
            ws.cell(row=row_idx, column=5, value=land_info.get("zone_type", "N/A"))
            ws.cell(row=row_idx, column=6, value=lh_review.get("lh_score_total", 0))
            ws.cell(row=row_idx, column=7, value=lh_review.get("judgement", "N/A"))
            ws.cell(row=row_idx, column=8, value=lh_review.get("grade", "N/A"))
            ws.cell(row=row_idx, column=9, value=feasibility.get("npv", 0))
            ws.cell(row=row_idx, column=10, value=feasibility.get("irr", 0))
            
            # 스타일
            for col in range(1, 11):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = self.normal_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 10
    
    def _create_financial_analysis_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """재무 분석 시트 생성"""
        ws = wb.create_sheet("재무 분석")
        
        # 헤더
        headers = [
            "부지 ID", "주소", "토지비", "건축비", "총사업비",
            "LH매입수익", "총수익", "NPV", "IRR(%)", "수익성"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # 데이터 입력
        sites = data.get("sites_analyzed", [])
        
        for row_idx, site in enumerate(sites, 2):
            land_info = site.get("land_info", {})
            feasibility = site.get("feasibility", {})
            
            ws.cell(row=row_idx, column=1, value=site.get("site_id", "N/A"))
            ws.cell(row=row_idx, column=2, value=land_info.get("address", "N/A"))
            ws.cell(row=row_idx, column=3, value=feasibility.get("land_cost", 0))
            ws.cell(row=row_idx, column=4, value=feasibility.get("construction_cost", 0))
            ws.cell(row=row_idx, column=5, value=feasibility.get("total_cost", 0))
            ws.cell(row=row_idx, column=6, value=feasibility.get("lh_revenue", 0))
            ws.cell(row=row_idx, column=7, value=feasibility.get("total_revenue", 0))
            ws.cell(row=row_idx, column=8, value=feasibility.get("npv", 0))
            ws.cell(row=row_idx, column=9, value=feasibility.get("irr", 0))
            ws.cell(row=row_idx, column=10, value=feasibility.get("profitability_grade", "N/A"))
            
            # 스타일
            for col in range(1, 11):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = self.normal_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 열 너비 조정
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['B'].width = 40
    
    def _create_lh_evaluation_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """LH 평가 시트 생성"""
        ws = wb.create_sheet("LH 평가")
        
        # 헤더
        headers = [
            "부지 ID", "주소", "A. 정책", "B. 입지", "C. 건설",
            "D. 가격", "E. 사업성", "총점", "등급", "판정"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # 데이터 입력
        sites = data.get("sites_analyzed", [])
        
        for row_idx, site in enumerate(sites, 2):
            land_info = site.get("land_info", {})
            lh_review = site.get("lh_review", {})
            sections = lh_review.get("section_scores", {})
            
            ws.cell(row=row_idx, column=1, value=site.get("site_id", "N/A"))
            ws.cell(row=row_idx, column=2, value=land_info.get("address", "N/A"))
            ws.cell(row=row_idx, column=3, value=sections.get("A", 0))
            ws.cell(row=row_idx, column=4, value=sections.get("B", 0))
            ws.cell(row=row_idx, column=5, value=sections.get("C", 0))
            ws.cell(row=row_idx, column=6, value=sections.get("D", 0))
            ws.cell(row=row_idx, column=7, value=sections.get("E", 0))
            ws.cell(row=row_idx, column=8, value=lh_review.get("lh_score_total", 0))
            ws.cell(row=row_idx, column=9, value=lh_review.get("grade", "N/A"))
            ws.cell(row=row_idx, column=10, value=lh_review.get("judgement", "N/A"))
            
            # 스타일
            for col in range(1, 11):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = self.normal_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        for col in range(3, 11):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    def _create_ranking_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """추천 순위 시트 생성"""
        ws = wb.create_sheet("추천 순위")
        
        # 제목
        ws.merge_cells('A1:E1')
        ws['A1'] = "부지 추천 순위 (LH 점수 기준)"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 헤더
        headers = ["순위", "부지 ID", "주소", "LH 점수", "판정"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # 데이터 입력 (이미 정렬된 상태)
        ranking = data.get("ranking_by_lh_score", [])
        
        for row_idx, site in enumerate(ranking, 4):
            ws.cell(row=row_idx, column=1, value=site.get("rank", row_idx-3))
            ws.cell(row=row_idx, column=2, value=site.get("site_id", "N/A"))
            ws.cell(row=row_idx, column=3, value=site.get("address", "N/A"))
            ws.cell(row=row_idx, column=4, value=site.get("lh_score", 0))
            ws.cell(row=row_idx, column=5, value=site.get("judgement", "N/A"))
            
            # 스타일
            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = self.normal_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        
        # 최종 추천
        final_row = len(ranking) + 6
        ws.merge_cells(f'A{final_row}:E{final_row}')
        
        final_recommendation = data.get("final_recommendation", {})
        ws[f'A{final_row}'] = f"최종 추천: {final_recommendation.get('site_id', 'N/A')}"
        ws[f'A{final_row}'].font = Font(name='맑은 고딕', size=12, bold=True, color="FF0000")
        ws[f'A{final_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    def _format_currency(self, value: float) -> str:
        """통화 포맷팅"""
        if value >= 1e8:
            return f"{value/1e8:.1f}억"
        elif value >= 1e4:
            return f"{value/1e4:.0f}만"
        else:
            return f"{value:,.0f}"


# 테스트 코드
if __name__ == "__main__":
    generator = ExcelComparisonReportGenerator()
    
    # 예제 데이터
    example_data = {
        "report_id": "M8-COMPARISON-20251227",
        "generated_at": "2025-12-27 00:00:00",
        "sites_analyzed": [],
        "summary_statistics": {
            "average_lh_score": 69.3,
            "average_npv": 937566666,
            "average_irr": 7.65,
            "go_count": 0,
            "conditional_go_count": 0,
            "no_go_count": 3
        }
    }
    
    report_path = generator.generate_comparison_report(example_data)
    print(f"엑셀 보고서 생성: {report_path}")
