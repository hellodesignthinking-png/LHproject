"""
ZeroSite v4.0 Attachment Manager
=================================

첨부 서류 생성 및 패키징

Author: ZeroSite M9 Team
Date: 2025-12-26
"""

import os
import json
import zipfile
from typing import Dict, Any, List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class AttachmentManager:
    """첨부 서류 생성 및 관리"""
    
    def __init__(self, output_dir: str = "output/proposals"):
        """초기화"""
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
        self.attachments = []
    
    def create_site_info_sheet(
        self,
        file_name: str,
        land_ctx: Any,
        m2_result: Any
    ) -> str:
        """부지 정보 시트 생성 (Excel)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "부지 정보"
        
        # 스타일 정의
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더
        ws['A1'] = "항목"
        ws['B1'] = "내용"
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['B1'].fill = header_fill
        ws['B1'].font = header_font
        
        # 데이터
        data = [
            ["주소", land_ctx.address],
            ["필지번호", land_ctx.parcel_id],
            ["면적 (㎡)", f"{land_ctx.area_sqm:,.2f}"],
            ["면적 (평)", f"{land_ctx.area_pyeong:,.2f}"],
            ["용도지역", land_ctx.zone_type],
            ["용적률", f"{land_ctx.far}%"],
            ["건폐율", f"{land_ctx.bcr}%"],
            ["도로폭", f"{land_ctx.road_width}m"],
            ["감정평가액", f"₩{m2_result.land_value:,}"],
            ["평당 단가", f"₩{m2_result.unit_price_pyeong:,}"],
            ["㎡당 단가", f"₩{m2_result.unit_price_sqm:,}"]
        ]
        
        for i, (key, value) in enumerate(data, start=2):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].border = border
            ws[f'B{i}'].border = border
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        
        # 저장
        file_path = os.path.join(self.output_dir, file_name)
        wb.save(file_path)
        
        print(f"✓ Site info sheet created: {file_path}")
        self.attachments.append(file_path)
        
        return file_path
    
    def create_financial_sheet(
        self,
        file_name: str,
        m5_result: Any
    ) -> str:
        """재무 분석 시트 생성 (Excel)"""
        wb = Workbook()
        
        # 비용 시트
        ws_cost = wb.active
        ws_cost.title = "비용 분석"
        
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # 비용 항목
        ws_cost['A1'] = "비용 항목"
        ws_cost['B1'] = "금액 (원)"
        ws_cost['A1'].fill = header_fill
        ws_cost['A1'].font = header_font
        ws_cost['B1'].fill = header_fill
        ws_cost['B1'].font = header_font
        
        cost_data = [
            ["토지매입비", m5_result.cost_breakdown.land_acquisition_cost],
            ["건축비", m5_result.cost_breakdown.construction_cost],
            ["설계비", m5_result.cost_breakdown.design_cost],
            ["간접비", m5_result.cost_breakdown.indirect_cost],
            ["금융비용", m5_result.cost_breakdown.financing_cost],
            ["예비비", m5_result.cost_breakdown.contingency],
            ["총 사업비", m5_result.cost_breakdown.total_cost]
        ]
        
        for i, (key, value) in enumerate(cost_data, start=2):
            ws_cost[f'A{i}'] = key
            ws_cost[f'B{i}'] = f"₩{value:,.0f}"
        
        ws_cost.column_dimensions['A'].width = 20
        ws_cost.column_dimensions['B'].width = 25
        
        # 수익 시트
        ws_revenue = wb.create_sheet(title="수익 분석")
        
        ws_revenue['A1'] = "수익 항목"
        ws_revenue['B1'] = "금액 (원)"
        ws_revenue['A1'].fill = header_fill
        ws_revenue['A1'].font = header_font
        ws_revenue['B1'].fill = header_fill
        ws_revenue['B1'].font = header_font
        
        revenue_data = [
            ["LH 매입가", m5_result.revenue_projection.lh_purchase_price],
            ["민간 분양 수익", m5_result.revenue_projection.private_sale_revenue],
            ["연간 임대 수익", m5_result.revenue_projection.rental_income_annual],
            ["총 수익", m5_result.revenue_projection.total_revenue]
        ]
        
        for i, (key, value) in enumerate(revenue_data, start=2):
            ws_revenue[f'A{i}'] = key
            ws_revenue[f'B{i}'] = f"₩{value:,.0f}"
        
        ws_revenue.column_dimensions['A'].width = 20
        ws_revenue.column_dimensions['B'].width = 25
        
        # 수익성 지표 시트
        ws_metrics = wb.create_sheet(title="수익성 지표")
        
        ws_metrics['A1'] = "지표"
        ws_metrics['B1'] = "값"
        ws_metrics['A1'].fill = header_fill
        ws_metrics['A1'].font = header_font
        ws_metrics['B1'].fill = header_fill
        ws_metrics['B1'].font = header_font
        
        metrics_data = [
            ["NPV (공공)", f"₩{m5_result.financial_metrics.npv_public:,.0f}"],
            ["IRR (공공)", f"{m5_result.financial_metrics.irr_public:.2f}%"],
            ["수익성 등급", m5_result.profitability_grade]
        ]
        
        for i, (key, value) in enumerate(metrics_data, start=2):
            ws_metrics[f'A{i}'] = key
            ws_metrics[f'B{i}'] = value
        
        ws_metrics.column_dimensions['A'].width = 20
        ws_metrics.column_dimensions['B'].width = 25
        
        # 저장
        file_path = os.path.join(self.output_dir, file_name)
        wb.save(file_path)
        
        print(f"✓ Financial sheet created: {file_path}")
        self.attachments.append(file_path)
        
        return file_path
    
    def create_capacity_sheet(
        self,
        file_name: str,
        m4_result: Any
    ) -> str:
        """건축 규모 시트 생성 (Excel)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "건축 규모"
        
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # 헤더
        ws['A1'] = "구분"
        ws['B1'] = "법정 용적률"
        ws['C1'] = "인센티브 용적률"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}1'].fill = header_fill
            ws[f'{col}1'].font = header_font
            ws[f'{col}1'].alignment = Alignment(horizontal='center')
        
        # 데이터
        legal = m4_result.legal_capacity
        incentive = m4_result.incentive_capacity
        
        data = [
            ["용적률", f"{m4_result.input_legal_far}%", f"{m4_result.input_incentive_far}%"],
            ["건폐율", f"{legal.applied_bcr}%", f"{incentive.applied_bcr}%"],
            ["연면적", f"{legal.target_gfa_sqm:,.1f}m²", f"{incentive.target_gfa_sqm:,.1f}m²"],
            ["세대수", f"{legal.total_units}세대", f"{incentive.total_units}세대"],
            ["주차대수", f"{legal.required_parking_spaces}대", f"{incentive.required_parking_spaces}대"]
        ]
        
        for i, row_data in enumerate(data, start=2):
            ws[f'A{i}'] = row_data[0]
            ws[f'B{i}'] = row_data[1]
            ws[f'C{i}'] = row_data[2]
        
        # 열 너비
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        # 저장
        file_path = os.path.join(self.output_dir, file_name)
        wb.save(file_path)
        
        print(f"✓ Capacity sheet created: {file_path}")
        self.attachments.append(file_path)
        
        return file_path
    
    def create_m6_report_json(
        self,
        file_name: str,
        m6_result: Any
    ) -> str:
        """M6 판정 결과 JSON 생성"""
        data = m6_result.to_dict()
        
        file_path = os.path.join(self.output_dir, file_name)
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"✓ M6 report JSON created: {file_path}")
        self.attachments.append(file_path)
        
        return file_path
    
    def create_submission_package(
        self,
        package_name: str,
        main_doc_path: str,
        include_source: bool = True
    ) -> str:
        """제출 패키지 ZIP 생성"""
        zip_path = os.path.join(self.output_dir, package_name)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # 주 문서 추가
            if os.path.exists(main_doc_path):
                zipf.write(main_doc_path, os.path.basename(main_doc_path))
            
            # 첨부 파일 추가
            for attachment in self.attachments:
                if os.path.exists(attachment):
                    zipf.write(attachment, os.path.basename(attachment))
        
        print(f"✓ Submission package created: {zip_path}")
        print(f"  - Main document: {os.path.basename(main_doc_path)}")
        print(f"  - Attachments: {len(self.attachments)} files")
        
        return zip_path
    
    def get_attachments(self) -> List[str]:
        """첨부 파일 목록 반환"""
        return self.attachments.copy()


__all__ = ['AttachmentManager']
